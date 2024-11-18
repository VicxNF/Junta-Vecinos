from django.shortcuts import render, redirect, get_object_or_404
from django.core.mail import send_mail, EmailMessage
from django.conf import settings
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login as auth_login, logout 
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from .models import *
from .forms import *
import uuid
from django.utils import timezone
from django.db import IntegrityError
from django.http import HttpResponseBadRequest, JsonResponse, HttpResponse
from django.utils.dateparse import parse_time
import json
from django.template.loader import render_to_string
from django.http import HttpResponse
from weasyprint import HTML
from django.core.files.base import ContentFile
from datetime import datetime, date
from django.core.files.base import ContentFile
import json
from datetime import datetime, timedelta
from django.db.models import Count, Sum, F
from reportlab.lib.pagesizes import A4, letter
from reportlab.pdfgen import canvas
from django.shortcuts import redirect
from django.urls import reverse
from transbank.webpay.webpay_plus.transaction import Transaction
from django.views.decorators.csrf import csrf_exempt
from decimal import Decimal
from django.utils.crypto import get_random_string
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from django.db.models.functions import TruncMonth, ExtractMonth
from calendar import month_name
import calendar
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter



def bienvenida(request):
    if request.user.is_authenticated:
        return redirect('index')
    
    # Obtener noticias ordenadas por fecha de publicación
    noticias = Noticia.objects.all().order_by('-fecha_publicacion')[:5]
    
    context = {
        'noticias': noticias,
        'espacios': Espacio.objects.all()[:3],
        'actividades': ActividadVecinal.objects.filter(estado='activa').order_by('fecha')[:3],
        'total_vecinos': Vecino.objects.count(),
        'total_proyectos': ProyectoVecinal.objects.filter(estado='aprobado').count(),
        'total_actividades': ActividadVecinal.objects.filter(estado='activa').count(),
        'total_espacios': Espacio.objects.count(),
    }
    
    return render(request, 'junta_vecinos/bienvenida.html', context)

@login_required
def index(request):
    vecinos = []
    solicitudes = []
    postulaciones = []
    noticias = []
    espacios = []
    reservas = []
    comuna = ""

    if request.user.is_authenticated:
        if request.user.is_superuser:
            admin = AdministradorComuna.objects.get(user=request.user)
            comuna = admin.get_comuna_display()
            
            vecinos = Vecino.objects.filter(administrador=admin)[:5]
            solicitudes = SolicitudCertificado.objects.filter(vecino__administrador=admin)[:5]
            postulaciones = ProyectoVecinal.objects.filter(vecino__administrador=admin)[:5]
            espacios = Espacio.objects.filter(comuna=admin)[:5]
            reservas = Reserva.objects.filter(espacio__comuna=admin)[:5]
            noticias = Noticia.objects.filter(comuna=admin).order_by('-fecha_publicacion')[:5]
        else:
            try:
                vecino = Vecino.objects.get(user=request.user)
                comuna = vecino.get_comuna_display()
                noticias = Noticia.objects.filter(comuna__comuna=vecino.comuna).order_by('-fecha_publicacion')[:5]
            except Vecino.DoesNotExist:
                noticias = []

    return render(request, 'junta_vecinos/index.html', {
        'vecinos': vecinos,
        'solicitudes': solicitudes,
        'postulaciones': postulaciones,
        'noticias': noticias,
        'espacios': espacios,
        'reservas': reservas,
        'comuna': comuna
    })


def detalle_espacio(request, espacio_id):
    espacio = get_object_or_404(Espacio, id=espacio_id)
    return render(request, 'junta_vecinos/detalle_espacio.html', {'espacio': espacio})

def detalle_actividad(request, actividad_id):
    actividad = get_object_or_404(ActividadVecinal, id=actividad_id)
    return render(request, 'junta_vecinos/detalle_actividad.html', {'actividad': actividad})

def is_admin(user):
    return hasattr(user, 'administradorcomuna')  # Solo permite el acceso si el usuario es un superusuario (admin)

def generar_username_unico(base_username):
    username = base_username
    n = 1
    while User.objects.filter(username=username).exists():
        username = f"{base_username}{n}"
        n += 1
    return username

@user_passes_test(is_admin)
@login_required
def lista_vecinos(request):
    admin = request.user.administradorcomuna
    vecinos = Vecino.objects.filter(comuna=admin.comuna)
    return render(request, 'junta_vecinos/lista_vecinos.html', {'vecinos': vecinos})

def registro_vecino(request):
    if request.method == 'POST':
        form = RegistroVecinoForm(request.POST)
        if form.is_valid():
            vecino = form.save()
            messages.success(request, 'Su solicitud de registro ha sido enviada y está pendiente de aprobación.')
            return redirect('login')
    else:
        form = RegistroVecinoForm()
    return render(request, 'junta_vecinos/registro.html', {'form': form})

def user_login(request):
    if request.method == "POST":
        form = LoginForm(request.POST)
        email = request.POST.get('email')
        password = request.POST.get('password')
        tipo_usuario = request.POST.get('tipo_usuario')

        try:
            user = User.objects.get(email=email)
            
            # Primero verificamos si el usuario está bloqueado
            if LoginAttempt.is_user_locked(user):
                messages.error(
                    request,
                    "Tu cuenta está temporalmente bloqueada por múltiples intentos fallidos. "
                    "Por favor, intenta nuevamente en 30 minutos."
                )
                return render(request, 'junta_vecinos/login.html', {'form': form})

            # Intentamos autenticar al usuario
            user_auth = authenticate(request, username=user.username, password=password)
            
            # Registramos el intento independientemente del resultado
            login_attempt = LoginAttempt.objects.create(
                user=user,
                ip_address=request.META.get('REMOTE_ADDR'),
                was_successful=user_auth is not None
            )

            if user_auth is not None:
                # Verificar el tipo de usuario
                if tipo_usuario == 'administrador' and hasattr(user, 'administradorcomuna'):
                    if user.is_active:
                        auth_login(request, user)
                        messages.success(request, f"¡Bienvenido Administrador {user.get_full_name()}!")
                        return redirect('index')
                elif tipo_usuario == 'vecino' and hasattr(user, 'vecino'):
                    if user.is_active:
                        auth_login(request, user)
                        messages.success(request, f"¡Bienvenido Vecino {user.get_full_name()}!")
                        return redirect('index')
                    else:
                        messages.error(request, "Su cuenta aún no ha sido aprobada.")
                else:
                    messages.error(request, "Tipo de usuario incorrecto.")
            else:
                # Verificar intentos restantes
                attempts_left = LoginAttempt.get_attempts_left(user)
                if attempts_left > 0:
                    messages.error(
                        request,
                        f"Correo electrónico o contraseña incorrectos. "
                        f"Te quedan {attempts_left} {'intento' if attempts_left == 1 else 'intentos'}."
                    )
                else:
                    messages.error(
                        request,
                        "Has excedido el número máximo de intentos. "
                        "Tu cuenta ha sido bloqueada temporalmente por 30 minutos."
                    )

        except User.DoesNotExist:
            messages.error(request, "El correo electrónico no está registrado.")
    else:
        form = LoginForm()
    
    return render(request, 'junta_vecinos/login.html', {'form': form})


def user_logout(request):
    logout(request)
    messages.success(request, "Has cerrado sesión exitosamente.")
    return redirect('login')

def solicitar_reestablecer_contrasena(request):
    if request.method == 'POST':
        formulario = FormularioSolicitudReestablecerContrasena(request.POST)
        if formulario.is_valid():
            correo = formulario.cleaned_data['correo']
            usuario = User.objects.get(email=correo)
            
            # Generar token único
            token = get_random_string(length=32)
            
            # Guardar token en la base de datos
            TokenReestablecerContrasena.objects.create(
                usuario=usuario,
                token=token,
                fecha_expiracion=timezone.now() + timezone.timedelta(hours=24)
            )
            
            # Construir el enlace de restablecimiento
            enlace_reestablecimiento = request.build_absolute_uri(
                reverse('confirmar_reestablecimiento_contrasena', kwargs={'token': token})
            )
            
            # Enviar correo
            send_mail(
                'Reestablecimiento de Contraseña',
                f'Para reestablecer tu contraseña, haz clic en el siguiente enlace:\n\n'
                f'{enlace_reestablecimiento}\n\n'
                f'Este enlace expirará en 24 horas.',
                settings.DEFAULT_FROM_EMAIL,
                [correo],
                fail_silently=False,
            )
            
            messages.success(
                request,
                "Te hemos enviado un correo electrónico con instrucciones para reestablecer tu contraseña."
            )
            return redirect('login')
    else:
        formulario = FormularioSolicitudReestablecerContrasena()
    
    return render(request, 'junta_vecinos/solicitar_reestablecer_contrasena.html', 
                 {'formulario': formulario})

def confirmar_reestablecimiento_contrasena(request, token):
    try:
        token_reestablecimiento = TokenReestablecerContrasena.objects.get(
            token=token,
            usado=False,
            fecha_expiracion__gt=timezone.now()
        )
    except TokenReestablecerContrasena.DoesNotExist:
        messages.error(request, "El enlace de restablecimiento es inválido o ha expirado.")
        return redirect('login')
    
    if request.method == 'POST':
        formulario = FormularioNuevaContrasena(request.POST)
        if formulario.is_valid():
            usuario = token_reestablecimiento.usuario
            usuario.set_password(formulario.cleaned_data['contrasena1'])
            usuario.save()
            
            # Marcar token como usado
            token_reestablecimiento.usado = True
            token_reestablecimiento.save()
            
            messages.success(request, "Tu contraseña ha sido actualizada correctamente.")
            return redirect('login')
    else:
        formulario = FormularioNuevaContrasena()
    
    return render(request, 'junta_vecinos/confirmar_reestablecimiento_contrasena.html', 
                 {'formulario': formulario})

@login_required
def solicitudes_registro(request):
    if not hasattr(request.user, 'administradorcomuna'):
        messages.error(request, 'No tienes permiso para acceder a esta página.')
        return redirect('index')
    
    admin = request.user.administradorcomuna
    solicitudes = SolicitudRegistroVecino.objects.filter(
        vecino__comuna=admin.comuna,
        is_approved=False
    )
    return render(request, 'junta_vecinos/solicitudes_registro.html', {'solicitudes': solicitudes})

@login_required
def aprobar_registro(request, solicitud_id):
    if not hasattr(request.user, 'administradorcomuna'):
        messages.error(request, 'No tienes permiso para realizar esta acción.')
        return redirect('index')
    
    solicitud = get_object_or_404(SolicitudRegistroVecino, id=solicitud_id)
    
    # Verificar que el administrador pertenece a la misma comuna que el vecino
    if solicitud.vecino.comuna != request.user.administradorcomuna.comuna:
        messages.error(request, 'No tienes permiso para aprobar esta solicitud.')
        return redirect('solicitudes_registro')
    
    vecino = solicitud.vecino
    user = vecino.user
    user.is_active = True
    user.save()
    
    solicitud.is_approved = True
    solicitud.save()

    send_mail(
        'Solicitud de Registro Aprobada',
        f'Hola {vecino.nombres},\n\nNos complace informarte que tu solicitud de registro ha sido aprobada. Ya puedes ingresar al sistema.\n\nSaludos!',
        settings.DEFAULT_FROM_EMAIL,
        [user.email],
    )

    messages.success(request, f'La solicitud de {vecino.nombres} {vecino.apellidos} ha sido aprobada y se le ha enviado un correo.')
    return redirect('solicitudes_registro')

@login_required
def rechazar_registro(request, solicitud_id):
    if not hasattr(request.user, 'administradorcomuna'):
        messages.error(request, 'No tienes permiso para realizar esta acción.')
        return redirect('index')

    solicitud = get_object_or_404(SolicitudRegistroVecino, id=solicitud_id)
    
    # Verificar que el administrador pertenece a la misma comuna que el vecino
    if solicitud.vecino.comuna != request.user.administradorcomuna.comuna:
        messages.error(request, 'No tienes permiso para rechazar esta solicitud.')
        return redirect('solicitudes_registro')

    vecino = solicitud.vecino
    user = vecino.user

    if request.method == 'POST':
        form = RechazoCertificadoForm(request.POST)
        if form.is_valid():
            mensaje_rechazo = form.cleaned_data['mensaje_rechazo']
            
            # Eliminar al usuario y sus datos
            user.delete()
            vecino.delete()
            solicitud.delete()

            send_mail(
                'Solicitud de Registro Rechazada',
                f'Hola {vecino.nombres} {vecino.apellidos},\n\n'
                f'Lamentamos informarte que tu solicitud de registro ha sido rechazada por las siguientes razones:\n\n{mensaje_rechazo}\n\n'
                f'Si tienes dudas, por favor contacta con la administración.',
                settings.DEFAULT_FROM_EMAIL,
                [user.email]
            )

            messages.error(request, 'La solicitud ha sido rechazada y se ha enviado un correo al vecino.')
            return redirect('solicitudes_registro')
    else:
        form = RechazoCertificadoForm()

    return render(request, 'junta_vecinos/rechazar_registro.html', {'solicitud': solicitud, 'form': form})


@login_required
def solicitar_certificado(request):
    try:
        vecino = request.user.vecino
    except Vecino.DoesNotExist:
        messages.error(request, 'No tienes un perfil de vecino asociado. Por favor, contacta con el administrador.')
        return redirect('index')

    if request.method == 'POST':
        form = SolicitudCertificadoForm(request.POST, request.FILES)
        if form.is_valid():
            solicitud = form.save(commit=False)
            solicitud.vecino = vecino
            solicitud.save()
            messages.success(request, 'Tu solicitud de certificado de residencia ha sido enviada.')
            return redirect('index')
    else:
        form = SolicitudCertificadoForm()

    return render(request, 'junta_vecinos/solicitar_certificado.html', {'form': form})

@user_passes_test(is_admin)
def gestionar_solicitudes(request):
    admin = request.user.administradorcomuna
    solicitudes = SolicitudCertificado.objects.filter(
        vecino__comuna=admin.comuna
    ).order_by('-fecha_solicitud')
    
    # Contar las solicitudes aprobadas y rechazadas
    aprobadas = solicitudes.filter(estado='Aprobado').count()
    rechazadas = solicitudes.filter(estado='Rechazado').count()
    
    return render(request, 'junta_vecinos/gestionar_solicitudes.html', {
        'solicitudes': solicitudes,
        'aprobadas': aprobadas,
        'rechazadas': rechazadas,
        'admin': admin
    })

@user_passes_test(is_admin)
@login_required
def generar_reporte_solicitudes_pdf(request):
    # Crear el objeto HttpResponse con el tipo de contenido PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="reporte_solicitudes.pdf"'

    # Crear el PDF
    p = canvas.Canvas(response, pagesize=A4)
    ancho, alto = A4

    # Título
    p.setFont("Helvetica-Bold", 16)
    p.drawString(100, alto - 50, "Reporte de Solicitudes de Certificados")

    # Encabezados de la tabla
    p.setFont("Helvetica-Bold", 12)
    p.drawString(50, alto - 100, "Vecino")
    p.drawString(200, alto - 100, "Fecha de Solicitud")
    p.drawString(350, alto - 100, "Estado")

    # Listar las solicitudes
    p.setFont("Helvetica", 10)
    y = alto - 120
    solicitudes = SolicitudCertificado.objects.all().order_by('fecha_solicitud')
    for solicitud in solicitudes:
        if y < 50:
            p.showPage()  # Crear nueva página si se acaba el espacio
            y = alto - 50

        p.drawString(50, y, f"{solicitud.vecino.nombres} {solicitud.vecino.apellidos}")
        p.drawString(200, y, solicitud.fecha_solicitud.strftime("%Y-%m-%d"))
        p.drawString(350, y, solicitud.get_estado_display())
        y -= 20

    # Finalizar el PDF
    p.showPage()
    p.save()
    
    return response

@user_passes_test(is_admin)
def ver_solicitud(request, id):
    solicitud = get_object_or_404(SolicitudCertificado, id=id)

    if request.method == 'POST' and 'aprobar' in request.POST:
        # Generar un número único de certificado
        numero_certificado = f"CERT-{solicitud.vecino.id}-{solicitud.id}"

        # Generar el PDF del certificado
        pdf_file = generar_certificado_pdf(solicitud.vecino, numero_certificado)

        # Guardar el certificado como un archivo
        certificado = CertificadoResidencia.objects.create(
            vecino=solicitud.vecino,
            numero_certificado=numero_certificado,
            documento_certificado=ContentFile(pdf_file, f"certificado_{numero_certificado}.pdf")
        )

        # Actualizar el estado de la solicitud
        solicitud.estado = 'Aprobado'
        solicitud.save()

        # Enviar el correo con el PDF adjunto
        email = EmailMessage(
            subject='Certificado de Residencia Aprobado',
            body=f'Hola {solicitud.vecino.nombres},\n\nEs de nuestro agrado informarle que su solicitud de certificado de residencia fue aprobada.\n\nAdjunto encontrarás tu certificado de residencia.',
            from_email=settings.DEFAULT_FROM_EMAIL,
            to=[solicitud.vecino.user.email],
        )

        # Adjuntar el PDF al correo
        email.attach(f"certificado_{numero_certificado}.pdf", pdf_file, 'application/pdf')

        # Enviar el correo
        email.send()

        messages.success(request, 'La solicitud ha sido aprobada, el certificado generado y enviado al vecino.')
        return redirect('gestionar_solicitudes')

    return render(request, 'junta_vecinos/ver_solicitud.html', {'solicitud': solicitud})


@user_passes_test(is_admin)
def rechazar_solicitud(request, id):
    solicitud = get_object_or_404(SolicitudCertificado, id=id)

    if request.method == 'POST':
        form = RechazoCertificadoForm(request.POST)
        if form.is_valid():
            mensaje_rechazo = form.cleaned_data['mensaje_rechazo']

            # Actualizar el estado de la solicitud
            solicitud.estado = 'Rechazado'
            solicitud.save()

            # Enviar el correo al vecino con las razones del rechazo
            send_mail(
                'Solicitud de Certificado de Residencia Rechazada',
                f'Hola {solicitud.vecino.user.get_full_name()},\n\n'
                f'Tu solicitud ha sido rechazada por las siguientes razones:\n\n{mensaje_rechazo}\n\n'
                f'Si tienes dudas, por favor contacta con la administración.',
                settings.DEFAULT_FROM_EMAIL,
                [solicitud.vecino.user.email]
            )

            messages.error(request, 'La solicitud ha sido rechazada y se ha enviado un correo al vecino.')
            return redirect('gestionar_solicitudes')
    else:
        form = RechazoCertificadoForm()

    return render(request, 'junta_vecinos/rechazar_solicitud.html', {'solicitud': solicitud, 'form': form})


@user_passes_test(is_admin)
def enviar_certificado(request, id):
    solicitud = get_object_or_404(SolicitudCertificado, id=id)

    if request.method == 'POST':
        form = EnviarCertificadoForm(request.POST, request.FILES)
        if form.is_valid():
            # Guardar el documento del certificado
            certificado = CertificadoResidencia.objects.create(
                vecino=solicitud.vecino,
                numero_certificado=f"CERT-{solicitud.vecino.id}-{solicitud.id}",
                documento_certificado=form.cleaned_data['documento_certificado']
            )
            
            # Enviar el correo
            send_mail(
                'Certificado de Residencia Aprobado',
                form.cleaned_data['contenido_correo'] + f'\n\nPuedes descargar el certificado desde el siguiente enlace:\n\n{settings.SITE_URL}/media/{certificado.documento_certificado}',
                settings.DEFAULT_FROM_EMAIL,
                [solicitud.vecino.user.email],
                fail_silently=False,
            )

            # Actualizar el estado de la solicitud
            solicitud.estado = 'Aprobado'
            solicitud.save()

            messages.success(request, 'El certificado ha sido enviado y la solicitud ha sido aprobada.')
            return redirect('gestionar_solicitudes')
    else:
        form = EnviarCertificadoForm()

    return render(request, 'junta_vecinos/enviar_certificado.html', {'form': form, 'solicitud': solicitud})

@login_required
def crear_proyecto(request):
    try:
        vecino = request.user.vecino
    except Vecino.DoesNotExist:
        messages.error(request, "Debes estar registrado como vecino para crear proyectos.")
        return redirect('index')

    if request.method == 'POST':
        form = ProyectoVecinalForm(request.POST, request.FILES)
        if form.is_valid():
            proyecto = form.save(commit=False)
            proyecto.vecino = vecino
            proyecto.save()
            
            # Enviar correo al administrador
            admin_email = vecino.administrador.user.email
            send_mail(
                'Nuevo Proyecto Vecinal',
                f'Se ha recibido un nuevo proyecto vecinal: {proyecto.propuesta}',
                settings.DEFAULT_FROM_EMAIL,
                [admin_email],
                fail_silently=True,
            )
            
            messages.success(request, "Tu proyecto ha sido enviado para revisión.")
            return redirect('mis_proyectos')
    else:
        form = ProyectoVecinalForm()
    
    return render(request, 'junta_vecinos/crear_proyecto.html', {'form': form})

@login_required
def mis_proyectos(request):
    try:
        vecino = request.user.vecino
        proyectos = ProyectoVecinal.objects.filter(vecino=vecino)
        
        # Calcular estadísticas
        total_proyectos = proyectos.count()
        proyectos_aprobados = proyectos.filter(estado='aprobado').count()
        proyectos_pendientes = proyectos.filter(estado='pendiente').count()
        proyectos_rechazados = proyectos.filter(estado='rechazado').count()
        
        context = {
            'proyectos': proyectos,
            'total_proyectos': total_proyectos,
            'proyectos_aprobados': proyectos_aprobados,
            'proyectos_pendientes': proyectos_pendientes,
            'proyectos_rechazados': proyectos_rechazados
        }
        
        return render(request, 'junta_vecinos/mis_proyectos.html', context)
    except Vecino.DoesNotExist:
        messages.error(request, "Debes estar registrado como vecino para ver proyectos.")
        return redirect('index')

@login_required
def detalle_proyecto(request, proyecto_id):
    proyecto = get_object_or_404(ProyectoVecinal, id=proyecto_id)
    
    # Primero verificamos si es un administrador
    if hasattr(request.user, 'administradorcomuna'):
        admin = request.user.administradorcomuna
        # Verificar si el proyecto pertenece a la comuna del administrador
        if proyecto.vecino.comuna == admin.comuna:
            return render(request, 'junta_vecinos/detalle_proyecto.html', {
                'proyecto': proyecto
            })
    
    # Si no es administrador, verificamos si es vecino
    try:
        vecino_solicitante = request.user.vecino
        
        # Verificar si es el creador del proyecto o un vecino de la misma comuna (proyecto aprobado)
        if (request.user == proyecto.vecino.user or 
            (vecino_solicitante.comuna == proyecto.vecino.comuna and 
             proyecto.estado == 'aprobado')):
            return render(request, 'junta_vecinos/detalle_proyecto.html', {
                'proyecto': proyecto
            })
        
    except Vecino.DoesNotExist:
        pass
    
    # Si no cumple ninguna condición, redirigir con mensaje de error
    messages.error(request, "No tienes permiso para ver este proyecto.")
    return redirect('index')

# Vista para el administrador
@login_required
def admin_proyectos(request):
    if not hasattr(request.user, 'administradorcomuna'):
        messages.error(request, "No tienes permiso para acceder a esta página.")
        return redirect('index')
    
    admin = request.user.administradorcomuna
    proyectos = ProyectoVecinal.objects.filter(vecino__comuna=admin.comuna)
    return render(request, 'junta_vecinos/admin_proyectos.html', {'proyectos': proyectos})

@login_required
def evaluar_proyecto(request, proyecto_id):
    if not hasattr(request.user, 'administradorcomuna'):
        messages.error(request, "No tienes permiso para realizar esta acción.")
        return redirect('index')
    
    proyecto = get_object_or_404(ProyectoVecinal, id=proyecto_id)
    
    if request.method == 'POST':
        estado = request.POST.get('estado')
        razon_rechazo = request.POST.get('razon_rechazo')
        
        proyecto.estado = estado
        if estado == 'rechazado':
            proyecto.razon_rechazo = razon_rechazo
        proyecto.save()
        
        # Enviar correo al vecino
        send_mail(
            'Actualización de tu Proyecto Vecinal',
            f'Tu proyecto "{proyecto.propuesta}" ha sido {estado}.' + 
            (f'\nRazón del rechazo: {razon_rechazo}' if estado == 'rechazado' else ''),
            settings.DEFAULT_FROM_EMAIL,
            [proyecto.vecino.user.email],
            fail_silently=True,
        )
        
        messages.success(request, f"El proyecto ha sido {estado}.")
        return redirect('admin_proyectos')
    
    return render(request, 'junta_vecinos/evaluar_proyecto.html', {'proyecto': proyecto})

@login_required
def proyectos_comuna(request):
    try:
        vecino = request.user.vecino
        proyectos = ProyectoVecinal.objects.filter(
            vecino__comuna=vecino.comuna,
            estado='aprobado'
        ).exclude(vecino=vecino)
        
        # Obtener las postulaciones del vecino actual
        postulaciones = {
            p.proyecto_id: p.estado 
            for p in PostulacionProyecto.objects.filter(vecino=vecino)
        }
        
        for proyecto in proyectos:
            proyecto.estado_postulacion = postulaciones.get(proyecto.id)
        
        return render(request, 'junta_vecinos/proyectos_comuna.html', {
            'proyectos': proyectos
        })
    except Vecino.DoesNotExist:
        messages.error(request, "Debes estar registrado como vecino para ver los proyectos.")
        return redirect('index')
    
@login_required
def postular_proyecto(request, proyecto_id):
    try:
        vecino = request.user.vecino
        proyecto = get_object_or_404(ProyectoVecinal, id=proyecto_id, estado='aprobado')
        
        # Verificar que el vecino sea de la misma comuna
        if vecino.comuna != proyecto.vecino.comuna:
            messages.error(request, "Solo puedes postular a proyectos de tu comuna.")
            return redirect('proyectos_comuna')
        
        # Verificar que no sea el creador del proyecto
        if proyecto.vecino == vecino:
            messages.error(request, "No puedes postular a tu propio proyecto.")
            return redirect('proyectos_comuna')
        
        # Verificar si ya existe una postulación
        if PostulacionProyecto.objects.filter(proyecto=proyecto, vecino=vecino).exists():
            messages.error(request, "Ya te has postulado a este proyecto.")
            return redirect('proyectos_comuna')
        
        if request.method == 'POST':
            form = PostulacionProyectoForm(request.POST)
            if form.is_valid():
                postulacion = form.save(commit=False)
                postulacion.proyecto = proyecto
                postulacion.vecino = vecino
                postulacion.save()
                
                # Enviar correo al creador del proyecto
                send_mail(
                    'Nueva postulación a tu proyecto',
                    f'El vecino {vecino.nombres} {vecino.apellidos} se ha postulado a tu proyecto "{proyecto.propuesta}".',
                    None,
                    [proyecto.vecino.user.email],
                    fail_silently=True,
                )
                
                messages.success(request, "Tu postulación ha sido enviada correctamente.")
                return redirect('proyectos_comuna')
        else:
            form = PostulacionProyectoForm()
        
        return render(request, 'junta_vecinos/postular_proyecto.html', {
            'form': form,
            'proyecto': proyecto
        })
        
    except Vecino.DoesNotExist:
        messages.error(request, "Debes estar registrado como vecino para postular.")
        return redirect('index')

@login_required
def gestionar_postulaciones(request, proyecto_id):
    proyecto = get_object_or_404(ProyectoVecinal, id=proyecto_id)
    
    # Verificar que el usuario sea el creador del proyecto
    if request.user != proyecto.vecino.user:
        messages.error(request, "No tienes permiso para gestionar las postulaciones de este proyecto.")
        return redirect('mis_proyectos')
    
    if request.method == 'POST':
        postulacion_id = request.POST.get('postulacion_id')
        accion = request.POST.get('accion')
        
        postulacion = get_object_or_404(PostulacionProyecto, id=postulacion_id, proyecto=proyecto)
        
        if accion in ['aceptada', 'rechazada']:
            postulacion.estado = accion
            postulacion.fecha_respuesta = timezone.now()
            postulacion.save()
            
            # Enviar correo al postulante
            estado_esp = 'aceptada' if accion == 'aceptada' else 'rechazada'
            send_mail(
                f'Tu postulación ha sido {estado_esp}',
                f'Tu postulación al proyecto "{proyecto.propuesta}" ha sido {estado_esp}.',
                None,
                [postulacion.vecino.user.email],
                fail_silently=True,
            )
            
            messages.success(request, f"La postulación ha sido {estado_esp}.")
        
    postulaciones = proyecto.postulaciones.all().order_by('-fecha_postulacion')
    return render(request, 'junta_vecinos/gestionar_postulaciones.html', {
        'proyecto': proyecto,
        'postulaciones': postulaciones
    })

@user_passes_test(is_admin)
def publicar_noticia(request):
    if request.method == 'POST':
        form = NoticiaForm(request.POST, request.FILES)
        if form.is_valid():
            noticia = form.save(commit=False)
            admin = AdministradorComuna.objects.get(user=request.user)
            noticia.comuna = admin
            noticia.save()
            return redirect('index')
    else:
        form = NoticiaForm()
    return render(request, 'junta_vecinos/publicar_noticia.html', {'form': form})


@login_required
@user_passes_test(is_admin)
def gestionar_noticias(request):
    admin = AdministradorComuna.objects.get(user=request.user)
    noticias = Noticia.objects.filter(comuna=admin).order_by('-fecha_publicacion')
    
    # Calcular estadísticas
    primer_dia_mes = timezone.now().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    noticias_mes = noticias.filter(fecha_publicacion__gte=primer_dia_mes).count()
    noticias_con_imagen = noticias.exclude(imagen='').count()
    
    context = {
        'noticias': noticias,
        'noticias_mes': noticias_mes,
        'noticias_con_imagen': noticias_con_imagen,
    }
    
    return render(request, 'junta_vecinos/gestionar_noticias.html', context)


@login_required
@user_passes_test(is_admin)
def editar_noticia(request, id):
    noticia = get_object_or_404(Noticia, id=id)
    if request.method == 'POST':
        form = NoticiaForm(request.POST, request.FILES, instance=noticia)
        if form.is_valid():
            form.save()
            return redirect('gestionar_noticias')
    else:
        form = NoticiaForm(instance=noticia)
    return render(request, 'junta_vecinos/editar_noticia.html', {'form': form, 'noticia': noticia})


def detalle_noticia(request, noticia_id):
    # Obtener la noticia o devolver 404 si no existe
    noticia = get_object_or_404(Noticia, id=noticia_id)
    
    # Obtener la comuna desde el contexto o configuración
    try:
        from django.conf import settings
        comuna = settings.COMUNA_NOMBRE
    except:
        comuna = "Comuna"

    # Contexto para la plantilla
    context = {
        'noticia': noticia,
        'comuna': noticia.comuna.get_comuna_display()  # Modificado para usar la comuna de la noticia
    }
    
    return render(request, 'junta_vecinos/detalle_noticia.html', context)


@user_passes_test(is_admin)
def registrar_espacio(request):
    if request.method == 'POST':
        form = EspacioForm(request.POST, request.FILES)
        if form.is_valid():
            espacio = form.save(commit=False)
            admin = AdministradorComuna.objects.get(user=request.user)
            espacio.comuna = admin
            espacio.save()
            return redirect('lista_espacios')  # Redirige a una página de éxito o lista de espacios
    else:
        form = EspacioForm()
    
    admin = AdministradorComuna.objects.get(user=request.user)
    context = {
        'form': form,
        'comuna': admin.get_comuna_display()
    }
    return render(request, 'junta_vecinos/registrar_espacio.html', context)

@user_passes_test(is_admin)
def lista_espacios(request):
    # Obtener la comuna del administrador autenticado
    administrador = AdministradorComuna.objects.get(user=request.user)
    
    # Filtrar los espacios según la comuna del administrador
    espacios = Espacio.objects.filter(comuna=administrador)
    
    return render(request, 'junta_vecinos/lista_espacios.html', {'espacios': espacios})

@user_passes_test(is_admin)
def editar_espacio(request, espacio_id):
    espacio = get_object_or_404(Espacio, id=espacio_id)
    
    if request.method == 'POST':
        form = EspacioForm(request.POST, instance=espacio)
        if form.is_valid():
            form.save()
            return redirect('lista_espacios')  # Redirige a la lista de espacios
    else:
        form = EspacioForm(instance=espacio)
    
    return render(request, 'junta_vecinos/editar_espacio.html', {'form': form, 'espacio': espacio})

@user_passes_test(is_admin)
def eliminar_espacio(request, espacio_id):
    espacio = get_object_or_404(Espacio, id=espacio_id)
    if request.method == 'POST':
        espacio.delete()
        return redirect('lista_espacios')  # Redirige a la lista después de eliminar
    return redirect('lista_espacios')  # Por si acceden a la URL por GET

@login_required
def espacios_disponibles(request):
    # Obtener el vecino asociado al usuario actual
    vecino = get_object_or_404(Vecino, user=request.user)
    
    # Filtrar los espacios según la comuna del vecino
    espacios = Espacio.objects.filter(comuna__comuna=vecino.comuna)
    
    espacio_seleccionado = None
    if request.method == 'GET' and 'espacio_id' in request.GET:
        espacio_id = request.GET.get('espacio_id')
        # Asegurarse de que el espacio seleccionado también pertenece a la comuna del vecino
        espacio_seleccionado = get_object_or_404(Espacio, id=espacio_id, comuna__comuna=vecino.comuna)
    
    return render(request, 'junta_vecinos/espacios_disponibles.html', {
        'espacios': espacios,
        'espacio_seleccionado': espacio_seleccionado
    })



@login_required
def get_espacio_info(request):
    espacio_id = request.GET.get('espacio_id')
    espacio = get_object_or_404(Espacio, id=espacio_id)
    return JsonResponse({
        'nombre': espacio.nombre,
        'ubicacion': espacio.ubicacion,
        'foto': espacio.foto.url if espacio.foto else '',
    })

@login_required
def reservar_espacio(request, espacio_id):
    espacio = get_object_or_404(Espacio, id=espacio_id)

    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            fecha = data.get('fecha')
            hora_inicio = data.get('hora_inicio')
            hora_fin = data.get('hora_fin')

            # Convertir las horas a objetos time
            hora_inicio_obj = datetime.strptime(hora_inicio, '%H:%M').time()
            hora_fin_obj = datetime.strptime(hora_fin, '%H:%M').time()

            # Verificar si ya existe una reserva para ese espacio, fecha y horas
            reservas_existentes = Reserva.objects.filter(
                espacio=espacio,
                fecha=fecha,
                hora_inicio__lt=hora_fin_obj,
                hora_fin__gt=hora_inicio_obj
            )
            if reservas_existentes.exists():
                return JsonResponse({'success': False, 'error': 'El espacio ya está reservado en parte o todo ese horario.'})

            # Calcular el monto a pagar (esto dependerá de tu lógica de negocio)
            monto = calcular_monto(espacio, hora_inicio_obj, hora_fin_obj)

            # Iniciar transacción WebPay
            buy_order = f"RES-{espacio.id}-{request.user.id}-{int(datetime.now().timestamp())}"
            session_id = request.session.session_key
            return_url = request.build_absolute_uri(reverse('webpay_retorno'))

            transaction = Transaction()
            response = transaction.create(buy_order, session_id, monto, return_url)

            # Guardar datos de la reserva en sesión para procesarla después del pago
            request.session['reserva_pendiente'] = {
                'espacio_id': espacio.id,
                'fecha': fecha,
                'hora_inicio': hora_inicio,
                'hora_fin': hora_fin,
                'monto': monto,
                'buy_order': buy_order
            }

            # Devolver la URL de pago y el token para que el frontend redirija
            return JsonResponse({
                'success': True,
                'payment_url': response['url'],
                'token': response['token']
            })
        
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})

    return JsonResponse({'success': False, 'error': 'Método no permitido'})

def calcular_monto(espacio, hora_inicio, hora_fin):
    # Calcular la duración en horas
    duracion = datetime.combine(datetime.min, hora_fin) - datetime.combine(datetime.min, hora_inicio)
    horas = Decimal(duracion.total_seconds()) / Decimal(3600)  # Convertimos a Decimal

    # Multiplicar el precio por hora (Decimal) con las horas (Decimal)
    return int(espacio.precio_por_hora * horas)

@csrf_exempt
def webpay_retorno(request):
    if request.method == 'POST' or request.method == 'GET':
        token = request.POST.get('token_ws') if request.method == 'POST' else request.GET.get('token_ws')
        
        if not token:
            return render(request, 'junta_vecinos/error_reserva.html', {'error': 'Token de pago no recibido'})
        
        transaction = Transaction()
        response = transaction.commit(token)

        if response['response_code'] == 0:
            # Pago exitoso
            reserva_data = request.session.get('reserva_pendiente')
            if reserva_data and reserva_data['buy_order'] == response['buy_order']:
                # Crear la reserva
                espacio = get_object_or_404(Espacio, id=reserva_data['espacio_id'])
                reserva = Reserva.objects.create(
                    usuario=request.user,
                    espacio=espacio,
                    fecha=reserva_data['fecha'],
                    hora_inicio=datetime.strptime(reserva_data['hora_inicio'], '%H:%M').time(),
                    hora_fin=datetime.strptime(reserva_data['hora_fin'], '%H:%M').time(),
                    monto_pagado=response['amount']
                )
                del request.session['reserva_pendiente']
                return render(request, 'junta_vecinos/reserva_exitosa.html', {'reserva': reserva})
            else:
                return render(request, 'junta_vecinos/error_reserva.html', {'error': 'Datos de reserva no coinciden'})
        else:
            # Pago fallido
            return render(request, 'junta_vecinos/pago_fallido.html')
    
    return HttpResponseBadRequest('Método no permitido')

# Vista para reserva exitosa
def reserva_exitosa(request):
    return render(request, 'junta_vecinos/reserva_exitosa.html')

# Vista para error en la reserva
def error_reserva(request):
    error = "Hubo un problema con tu reserva."  # Puedes pasar un mensaje de error personalizado
    return render(request, 'junta_vecinos/error_reserva.html', {'error': error})

# Vista para pago fallido
def pago_fallido(request):
    return render(request, 'junta_vecinos/pago_fallido.html')


@login_required
def get_available_slots(request):
    date = request.GET.get('date')
    espacio_id = request.GET.get('espacio_id')
    espacio = get_object_or_404(Espacio, id=espacio_id)

    # Definir horarios disponibles (ajusta según tus necesidades)
    start_time = datetime.strptime('08:00', '%H:%M').time()
    end_time = datetime.strptime('22:00', '%H:%M').time()
    slot_duration = timedelta(minutes=30)

    # Obtener todas las reservas para ese día y espacio
    reservas = Reserva.objects.filter(espacio=espacio, fecha=date)

    all_slots = []
    current_time = start_time
    while current_time < end_time:
        slot_end = (datetime.combine(datetime.min, current_time) + slot_duration).time()
        if slot_end > end_time:
            slot_end = end_time

        is_available = not reservas.filter(
            hora_inicio__lt=slot_end,
            hora_fin__gt=current_time
        ).exists()

        all_slots.append({
            'start': current_time.strftime('%H:%M'),
            'end': slot_end.strftime('%H:%M'),
            'available': is_available
        })

        current_time = slot_end

    return JsonResponse({'slots': all_slots})

@user_passes_test(is_admin)
@login_required
def lista_reservas(request):
    # Obtener el administrador y su comuna
    admin = AdministradorComuna.objects.get(user=request.user)
    
    # Obtener las reservas solo de los espacios de la comuna del administrador
    reservas = Reserva.objects.select_related('espacio', 'usuario').filter(
        espacio__comuna=admin
    ).order_by('-fecha', '-hora_inicio')

    # Obtener estadísticas por espacio de la comuna
    reservas_por_espacio = Reserva.objects.filter(
        espacio__comuna=admin
    ).values('espacio__nombre').annotate(
        total=Count('id'),
        ingresos_totales=Sum('monto_pagado')
    ).order_by('espacio__nombre')

    # Obtener estadísticas por mes para el año actual y la comuna específica
    year = datetime.now().year
    reservas_por_mes = Reserva.objects.filter(
        fecha__year=year,
        espacio__comuna=admin
    ).annotate(
        mes=TruncMonth('fecha')
    ).values('mes').annotate(
        total=Count('id'),
        ingresos=Sum('monto_pagado')
    ).order_by('mes')

    # Preparar datos para los gráficos
    espacios = [reserva['espacio__nombre'] for reserva in reservas_por_espacio]
    totales = [reserva['total'] for reserva in reservas_por_espacio]
    ingresos = [float(reserva['ingresos_totales'] or 0) for reserva in reservas_por_espacio]

    # Preparar datos para el gráfico de tendencia mensual
    meses = [calendar.month_name[r['mes'].month] for r in reservas_por_mes]
    reservas_mensuales = [r['total'] for r in reservas_por_mes]
    ingresos_mensuales = [float(r['ingresos'] or 0) for r in reservas_por_mes]

    # Obtener años para el selector de reportes
    current_year = datetime.now().year
    years = range(current_year - 2, current_year + 1)
    months = [(i, calendar.month_name[i]) for i in range(1, 13)]

    # Calcular el total de ingresos (para solucionar el problema del filtro sum)
    total_ingresos = sum(ingresos)

    return render(request, 'junta_vecinos/lista_reservas.html', {
        'reservas': reservas,
        'espacios': espacios,
        'totales': totales,
        'ingresos': ingresos,
        'total_ingresos': total_ingresos,
        'meses': meses,
        'reservas_mensuales': reservas_mensuales,
        'ingresos_mensuales': ingresos_mensuales,
        'years': years,
        'months': months,
        'current_year': current_year,
        'current_month': datetime.now().month,
        'comuna': admin.get_comuna_display(),  # Añadimos el nombre de la comuna para mostrarlo
    })


def generar_certificado_pdf(vecino, numero_certificado):
    # Renderizar el contenido HTML con los datos del vecino
    html_content = render_to_string('junta_vecinos/certificado_residencia.html', {
        'vecino': vecino,
        'numero_certificado': numero_certificado,
        'fecha_emision': date.today().strftime('%d/%m/%Y'),
        'junta_nombre': "Nombre de la Junta",  # Agregar el nombre real
        'comuna': vecino.comuna,
        'region': "Metropolitana",  # Agregar la región correspondiente
        'presidente_nombre': "Nombre del Presidente",  # Agregar el nombre real del presidente
        'presidente_ci': "12345678-9",  # Agregar el CI real del presidente
    })

    # Generar el PDF usando WeasyPrint
    pdf_file = HTML(string=html_content).write_pdf()

    return pdf_file

def get_month_name(month_number):
    return calendar.month_name[month_number]

@user_passes_test(is_admin)
@login_required
def generar_reporte_pdf(request):
    # Obtener parámetros del request
    year = int(request.GET.get('year', datetime.now().year))
    month = int(request.GET.get('month', datetime.now().month))
    
    # Crear el objeto HttpResponse con el tipo de contenido de PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="reporte_reservas_{year}_{month}.pdf"'
    
    # Configurar el documento
    doc = SimpleDocTemplate(response, pagesize=letter)
    elements = []
    styles = getSampleStyleSheet()
    
    # Estilos personalizados
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        spaceAfter=30,
        alignment=1  # Centrado
    )
    
    # Título del reporte
    title = Paragraph(f"Reporte de Reservas - {get_month_name(month)} {year}", title_style)
    elements.append(title)
    
    # Obtener las reservas del mes
    reservas = Reserva.objects.filter(
        fecha__year=year,
        fecha__month=month
    ).select_related('espacio', 'usuario')
    
    # Estadísticas generales
    total_reservas = reservas.count()
    ingresos_totales = reservas.aggregate(total=Sum('monto_pagado'))['total'] or 0
    
    # Agregar resumen general
    resumen_style = ParagraphStyle(
        'Resumen',
        parent=styles['Normal'],
        fontSize=12,
        spaceAfter=20
    )
    
    elements.append(Paragraph(f"Total de Reservas: {total_reservas}", resumen_style))
    elements.append(Paragraph(f"Ingresos Totales: ${ingresos_totales:,.2f}", resumen_style))
    elements.append(Spacer(1, 20))
    
    # Estadísticas por espacio
    stats_por_espacio = reservas.values('espacio__nombre').annotate(
        total_reservas=Count('id'),
        ingresos=Sum('monto_pagado')
    ).order_by('-total_reservas')
    
    # Tabla de estadísticas por espacio
    elements.append(Paragraph("Estadísticas por Espacio", styles['Heading2']))
    espacios_data = [['Espacio', 'Total Reservas', 'Ingresos']]
    for stat in stats_por_espacio:
        espacios_data.append([
            stat['espacio__nombre'],
            str(stat['total_reservas']),
            f"${stat['ingresos']:,.2f}" if stat['ingresos'] else "$0.00"
        ])
    
    tabla_espacios = Table(espacios_data)
    tabla_espacios.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('BOX', (0, 0), (-1, -1), 2, colors.black),
    ]))
    
    elements.append(tabla_espacios)
    elements.append(Spacer(1, 20))
    
    # Detalle de todas las reservas
    elements.append(Paragraph("Detalle de Reservas", styles['Heading2']))
    
    detalle_data = [['Fecha', 'Espacio', 'Usuario', 'Hora Inicio', 'Hora Fin', 'Monto']]
    for reserva in reservas.order_by('fecha', 'hora_inicio'):
        detalle_data.append([
            reserva.fecha.strftime("%d/%m/%Y"),
            reserva.espacio.nombre,
            f"{reserva.usuario.first_name} {reserva.usuario.last_name}",
            reserva.hora_inicio.strftime("%H:%M"),
            reserva.hora_fin.strftime("%H:%M"),
            f"${reserva.monto_pagado:,.2f}"
        ])
    
    tabla_detalle = Table(detalle_data, repeatRows=1)
    tabla_detalle.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('BOX', (0, 0), (-1, -1), 2, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    
    elements.append(tabla_detalle)
    
    # Generar el PDF
    doc.build(elements)
    return response

@login_required
def generar_reporte_actividades_excel(request):
    try:
        # Obtener parámetros con valores por defecto del año y mes actual
        year = request.GET.get('year', '')
        month = request.GET.get('month', '')
        
        # Si los parámetros están vacíos, usar fecha actual
        if not year or not month:
            current_date = datetime.now()
            year = current_date.year
            month = current_date.month
        else:
            year = int(year)
            month = int(month)
        
        admin = request.user.administradorcomuna
        
        # Filtrar actividades
        actividades = ActividadVecinal.objects.filter(
            comuna=admin,
            fecha__year=year,
            fecha__month=month
        ).order_by('fecha', 'hora_inicio')

        # Crear un nuevo libro de trabajo Excel
        wb = Workbook()
        ws = wb.active
        ws.title = f"Actividades {calendar.month_name[month]} {year}"

        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        total_font = Font(bold=True)
        total_fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Título del reporte
        ws['A1'] = f"Reporte de Actividades - {calendar.month_name[month]} {year}"
        ws.merge_cells('A1:D1')
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal="center")

        # Encabezados
        headers = ['Actividad', 'Fecha', 'Inscritos', 'Ingresos']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = border

        # Datos
        row = 4
        total_inscritos = 0
        total_ingresos = 0

        for actividad in actividades:
            inscritos = actividad.inscripcionactividad_set.count()
            ingresos = inscritos * actividad.precio
            
            ws.cell(row=row, column=1, value=actividad.titulo)
            ws.cell(row=row, column=2, value=actividad.fecha.strftime('%d/%m/%Y'))
            ws.cell(row=row, column=3, value=inscritos)
            ws.cell(row=row, column=4, value=float(ingresos))

            # Aplicar bordes y alineación
            for col in range(1, 5):
                cell = ws.cell(row=row, column=col)
                cell.border = border
                cell.alignment = Alignment(horizontal="center")

            total_inscritos += inscritos
            total_ingresos += ingresos
            row += 1

        # Totales
        row_total = row
        ws.cell(row=row_total, column=1, value="TOTAL")
        ws.cell(row=row_total, column=2, value="")
        ws.cell(row=row_total, column=3, value=total_inscritos)
        ws.cell(row=row_total, column=4, value=float(total_ingresos))

        # Aplicar estilos a la fila de totales
        for col in range(1, 5):
            cell = ws.cell(row=row_total, column=col)
            cell.font = total_font
            cell.fill = total_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center")

        # Formato de moneda para la columna de ingresos
        for row in range(4, row_total + 1):
            cell = ws.cell(row=row, column=4)
            cell.number_format = '"$"#,##0.00'

        # Ajustar el ancho de las columnas
        for col in range(1, 5):
            ws.column_dimensions[get_column_letter(col)].auto_size = True

        # Crear la respuesta HTTP con el archivo Excel
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="reporte_actividades_{year}_{month}.xlsx"'

        # Guardar el libro de trabajo
        wb.save(response)
        return response
        
    except Exception as e:
        # En caso de error, redirigir a la página de lista con un mensaje
        messages.error(request, f"Error al generar el reporte: {str(e)}")
        return redirect('lista_actividades')

# Vista para mostrar el formulario de selección de mes/año
@user_passes_test(is_admin)
@login_required
def seleccionar_periodo_reporte(request):
    current_year = datetime.now().year
    years = range(current_year - 2, current_year + 1)  # Últimos 2 años y año actual
    months = [(i, calendar.month_name[i]) for i in range(1, 13)]
    
    context = {
        'years': years,
        'months': months,
        'current_year': current_year,
        'current_month': datetime.now().month,
    }
    
    return render(request, 'junta_vecinos/seleccionar_periodo_reporte.html', context)

@user_passes_test(is_admin)
def crear_actividad(request):
    # Obtenemos directamente el AdministradorComuna
    admin = AdministradorComuna.objects.get(user=request.user)
    
    if request.method == 'POST':
        form = ActividadVecinalForm(request.POST, request.FILES, admin=admin)
        if form.is_valid():
            actividad = form.save(commit=False)
            actividad.comuna = admin
            actividad.estado = 'activa'
            
            # Obtenemos el nombre del espacio seleccionado
            espacio_seleccionado = form.cleaned_data['lugar']
            actividad.lugar = espacio_seleccionado.nombre
            
            actividad.save()
            
            # Notificar a todos los vecinos de la comuna
            vecinos = Vecino.objects.filter(comuna=admin.comuna)
            for vecino in vecinos:
                send_mail(
                    'Nueva Actividad Vecinal',
                    f'Se ha creado una nueva actividad: {actividad.titulo}\n'
                    f'Fecha: {actividad.fecha}\n'
                    f'Lugar: {actividad.lugar}\n'
                    f'Cupos disponibles: {actividad.cupo_maximo}',
                    settings.DEFAULT_FROM_EMAIL,
                    [vecino.user.email]
                )
            
            messages.success(request, 'Actividad creada exitosamente.')
            return redirect('lista_actividades')
    else:
        form = ActividadVecinalForm(admin=admin)
    
    return render(request, 'junta_vecinos/crear_actividad.html', {'form': form})

@login_required
def lista_actividades(request):
    current_year = datetime.now().year
    current_month = datetime.now().month

    if hasattr(request.user, 'administradorcomuna'):
        admin = request.user.administradorcomuna
        base_query = ActividadVecinal.objects.filter(comuna=admin)
        
        # Estadísticas generales
        total_actividades = base_query.count()
        total_inscritos = InscripcionActividad.objects.filter(
            actividad__comuna=admin
        ).count()
        total_ingresos = InscripcionActividad.objects.filter(
            actividad__comuna=admin
        ).aggregate(
            total=Sum(F('actividad__precio'))
        )['total'] or 0

        # Estadísticas por mes
        actividades_por_mes = base_query.filter(
            fecha__year=current_year
        ).annotate(
            mes=TruncMonth('fecha')
        ).values('mes').annotate(
            total=Count('id'),
            inscritos=Count('inscripcionactividad'),
            ingresos=Sum(F('precio') * F('cupo_actual'))
        ).order_by('mes')

        # Preparar datos para gráficos
        meses = [calendar.month_name[a['mes'].month] for a in actividades_por_mes]
        actividades_mensuales = [a['total'] for a in actividades_por_mes]
        inscritos_mensuales = [a['inscritos'] for a in actividades_por_mes]
        ingresos_mensuales = [float(a['ingresos'] or 0) for a in actividades_por_mes]

        # Datos para selector de reportes
        years = range(current_year - 2, current_year + 1)
        months = [(i, calendar.month_name[i]) for i in range(1, 13)]

        context = {
            'actividades': base_query.order_by('fecha', 'hora_inicio'),
            'total_actividades': total_actividades,
            'total_inscritos': total_inscritos,
            'total_ingresos': total_ingresos,
            'meses': meses,
            'actividades_mensuales': actividades_mensuales,
            'inscritos_mensuales': inscritos_mensuales,
            'ingresos_mensuales': ingresos_mensuales,
            'years': years,
            'months': months,
            'current_year': current_year,
            'current_month': current_month,
            'comuna': admin.get_comuna_display()
        }
    else:
        # Código para el vecino
        vecino = get_object_or_404(Vecino, user=request.user)
        actividades = ActividadVecinal.objects.filter(
            comuna__comuna=vecino.comuna,
            estado='activa'
        ).order_by('fecha', 'hora_inicio')

        # Verificar inscripción para cada actividad
        for actividad in actividades:
            actividad.usuario_inscrito = InscripcionActividad.objects.filter(
                actividad=actividad,
                vecino=vecino
            ).exists()
        
        context = {
            'actividades': actividades,
            'vecino': vecino  # Incluye el vecino en el contexto si se necesita para otro propósito
        }
    
    return render(request, 'junta_vecinos/lista_actividades.html', context)

@login_required
def generar_reporte_actividades_pdf(request):
    # Obtener parámetros
    year = int(request.GET.get('year', datetime.now().year))
    month = int(request.GET.get('month', datetime.now().month))
    
    admin = request.user.administradorcomuna
    
    # Filtrar actividades
    actividades = ActividadVecinal.objects.filter(
        comuna=admin,
        fecha__year=year,
        fecha__month=month
    ).order_by('fecha', 'hora_inicio')

    # Crear el PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="reporte_actividades_{year}_{month}.pdf"'
    
    doc = SimpleDocTemplate(response, pagesize=letter)
    elements = []
    styles = getSampleStyleSheet()
    
    # Título
    elements.append(Paragraph(f"Reporte de Actividades - {calendar.month_name[month]} {year}", styles['Title']))
    elements.append(Spacer(1, 20))
    
    # Datos de la tabla
    data = [['Actividad', 'Fecha', 'Inscritos', 'Ingresos']]
    total_inscritos = 0
    total_ingresos = 0
    
    for actividad in actividades:
        inscritos = actividad.inscripcionactividad_set.count()
        ingresos = inscritos * actividad.precio
        data.append([
            actividad.titulo,
            actividad.fecha.strftime('%d/%m/%Y'),
            str(inscritos),
            f"${ingresos:,.2f}"
        ])
        total_inscritos += inscritos
        total_ingresos += ingresos
    
    # Agregar totales
    data.append(['TOTAL', '', str(total_inscritos), f"${total_ingresos:,.2f}"])
    
    # Crear tabla
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 14),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
        ('TEXTCOLOR', (0, -1), (-1, -1), colors.black),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, -1), (-1, -1), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    elements.append(table)
    
    # Generar PDF
    doc.build(elements)
    return response

@login_required
def inscribir_actividad(request, actividad_id):
    actividad = get_object_or_404(ActividadVecinal, id=actividad_id)
    vecino = get_object_or_404(Vecino, user=request.user)
    
    if actividad.esta_llena():
        messages.error(request, 'Lo sentimos, esta actividad ya no tiene cupos disponibles.')
        return redirect('lista_actividades')
    
    if request.method == 'POST':
        try:
            
            # Crear transacción en WebPay
            buy_order = f"ACT-{actividad.id}-{vecino.user.id}-{int(datetime.now().timestamp())}"
            session_id = request.session.session_key
            return_url = request.build_absolute_uri(reverse('webpay_retorno_actividad'))
            monto = str(actividad.precio) 
            
            transaction = Transaction()
            response = transaction.create(buy_order, session_id, monto, return_url)

            # Guardar en la sesión los datos de inscripción
            request.session['inscripcion_pendiente'] = {
                'actividad_id': actividad.id,
                'monto': monto,
                'buy_order': buy_order
            }

            return JsonResponse({
                'success': True,
                'payment_url': response['url'],
                'token': response['token']
            })
        
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Método no permitido'})

@login_required
@user_passes_test(is_admin)
def eliminar_actividad(request, actividad_id):
    actividad = get_object_or_404(ActividadVecinal, id=actividad_id)

    # Verificar si la actividad se puede eliminar (por ejemplo, solo si está activa)
    if actividad.estado == 'activa':
        actividad.delete()
        messages.success(request, 'Actividad eliminada exitosamente.')
    else:
        messages.warning(request, 'No se puede eliminar esta actividad.')

    return redirect('lista_actividades')

@csrf_exempt
def webpay_retorno_actividad(request):
    if request.method == 'POST' or request.method == 'GET':
        token = request.POST.get('token_ws') if request.method == 'POST' else request.GET.get('token_ws')
        
        if not token:
            return render(request, 'junta_vecinos/error_inscripcion.html', {'error': 'Token de pago no recibido'})
        
        transaction = Transaction()
        response = transaction.commit(token)

        if response['response_code'] == 0:
            inscripcion_data = request.session.get('inscripcion_pendiente')
            if inscripcion_data and inscripcion_data['buy_order'] == response['buy_order']:
                # Crear la inscripción
                actividad = get_object_or_404(ActividadVecinal, id=inscripcion_data['actividad_id'])
                vecino = get_object_or_404(Vecino, user=request.user)
                InscripcionActividad.objects.create(actividad=actividad, vecino=vecino)
                actividad.cupo_actual += 1
                actividad.save()

                # Eliminar datos de inscripción de la sesión
                del request.session['inscripcion_pendiente']
                return render(request, 'junta_vecinos/inscripcion_exitosa.html', {'actividad': actividad})
            else:
                return render(request, 'junta_vecinos/error_inscripcion.html', {'error': 'Datos de inscripción no coinciden'})
        else:
            return render(request, 'junta_vecinos/pago_fallido.html')
    
    return HttpResponseBadRequest('Método no permitido')

def inscripcion_exitosa(request):
    actividad_id = request.session.get('inscripcion_pendiente', {}).get('actividad_id')
    actividad = get_object_or_404(ActividadVecinal, id=actividad_id)
    return render(request, 'junta_vecinos/inscripcion_exitosa.html', {'actividad': actividad})

def error_inscripcion(request):
    error_message = request.GET.get('error', 'Ha ocurrido un error desconocido.')
    return render(request, 'junta_vecinos/error_inscripcion.html', {'error': error_message})

@user_passes_test(is_admin)
def cancelar_actividad(request, actividad_id):
    actividad = get_object_or_404(ActividadVecinal, id=actividad_id)
    
    if request.method == 'POST':
        actividad.estado = 'cancelada'
        actividad.save()
        
        # Notificar a los inscritos
        inscripciones = InscripcionActividad.objects.filter(actividad=actividad)
        for inscripcion in inscripciones:
            send_mail(
                'Actividad Vecinal Cancelada',
                f'La actividad {actividad.titulo} ha sido cancelada.\n'
                f'Lamentamos los inconvenientes causados.',
                settings.DEFAULT_FROM_EMAIL,
                [inscripcion.vecino.user.email]
            )
        
        messages.success(request, 'Actividad cancelada exitosamente.')
        return redirect('lista_actividades')
    
    return render(request, 'junta_vecinos/cancelar_actividad.html', {
        'actividad': actividad
    })

@user_passes_test(is_admin)
def registrar_asistencia(request, actividad_id):
    actividad = get_object_or_404(ActividadVecinal, id=actividad_id)
    inscripciones = InscripcionActividad.objects.filter(actividad=actividad)
    
    if request.method == 'POST':
        asistentes = request.POST.getlist('asistentes')
        
        for inscripcion in inscripciones:
            inscripcion.asistio = str(inscripcion.id) in asistentes
            inscripcion.save()
        
        messages.success(request, 'Asistencia registrada exitosamente.')
        return redirect('lista_actividades')
    
    return render(request, 'junta_vecinos/registrar_asistencia.html', {
        'actividad': actividad,
        'inscripciones': inscripciones
    })